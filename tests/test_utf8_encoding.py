# test_utf8_encoding.py
import pytest
import json
import csv
import openpyxl
import myproject.saver as saver


# Тестируем кириллицу в сохраняемых файлах
@pytest.mark.encoding
def test_utf8_encoding(tmp_path):
    data = [{"name": "Тестовая строка", "url": "https://пример.ру"}]

    # CSV
    saver.OUTPUT_DIR = tmp_path
    saver.save_csv(data, filename="utf8.csv")
    csv_file = tmp_path / "utf8.csv"
    with csv_file.open(encoding="utf-8") as f:
        reader = list(csv.reader(f))
    # Проверяем заголовки (используются русские названия)
    assert "Название" in reader[0] and "Ссылка" in reader[0]
    # Проверяем данные (находим индексы колонок)
    name_idx = reader[0].index("Название")
    url_idx = reader[0].index("Ссылка")
    assert reader[1][name_idx] == "Тестовая строка"
    assert reader[1][url_idx] == "https://пример.ру"

    # JSON
    saver.save_json(data, filename="utf8.json")
    json_file = tmp_path / "utf8.json"
    parsed = json.loads(json_file.read_text(encoding="utf-8"))
    assert parsed[0]["name"] == "Тестовая строка"
    assert parsed[0]["url"] == "https://пример.ру"

    # XLSX
    saver.save_xlsx(data, filename="utf8.xlsx")
    xlsx_file = tmp_path / "utf8.xlsx"
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb.active
    assert sheet.cell(row=2, column=1).value == "Тестовая строка"
    assert sheet.cell(row=2, column=2).value == "https://пример.ру"

    # HTML
    saver.render_html(data, filename="utf8.html")
    html_file = tmp_path / "utf8.html"
    html_content = html_file.read_text(encoding="utf-8")
    assert "Тестовая строка" in html_content
    assert "https://пример.ру" in html_content
