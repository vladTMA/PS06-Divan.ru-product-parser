# test_saver.py
import pytest
import myproject.saver as saver
import json
import csv
import openpyxl


@pytest.mark.unit
def test_save_csv(tmp_path):
    data = [{"name": "Пример", "url": "https://example.com"}]
    saver.OUTPUT_DIR = tmp_path
    saver.save_csv(data, filename="test.csv")

    file_path = tmp_path / "test.csv"
    assert file_path.exists()

    # Проверяем содержимое CSV
    with file_path.open(encoding="utf-8") as f:
        reader = list(csv.reader(f))
    # Первая строка — заголовки (используются русские названия)
    assert "Название" in reader[0] and "Ссылка" in reader[0]
    # Вторая строка — данные (находим индексы колонок)
    name_idx = reader[0].index("Название")
    url_idx = reader[0].index("Ссылка")
    assert reader[1][name_idx] == "Пример"
    assert reader[1][url_idx] == "https://example.com"


@pytest.mark.unit
def test_save_json(tmp_path):
    data = [{"name": "Пример", "url": "https://example.com"}]
    saver.OUTPUT_DIR = tmp_path
    saver.save_json(data, filename="test.json")

    file_path = tmp_path / "test.json"
    assert file_path.exists()

    # Проверяем содержимое JSON
    content = json.loads(file_path.read_text(encoding="utf-8"))
    assert content[0]["name"] == "Пример"
    assert content[0]["url"] == "https://example.com"


@pytest.mark.unit
def test_save_xlsx(tmp_path):
    data = [{"name": "Пример", "url": "https://example.com"}]
    saver.OUTPUT_DIR = tmp_path
    saver.save_xlsx(data, filename="test.xlsx")

    file_path = tmp_path / "test.xlsx"
    assert file_path.exists()

    # Проверяем содержимое XLSX
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    # Используются русские названия заголовков
    assert "Название" in headers and "Ссылка" in headers
    # Находим индексы колонок
    name_idx = headers.index("Название") + 1  # openpyxl использует 1-based индексы
    url_idx = headers.index("Ссылка") + 1
    assert sheet.cell(row=2, column=name_idx).value == "Пример"
    assert sheet.cell(row=2, column=url_idx).value == "https://example.com"


@pytest.mark.unit
def test_render_html(tmp_path):
    data = [
        {
            "name": "Люстра",
            "price": "12990",
            "currency": "RUB",
            "url": "https://example.com/lamp",
            "instock_text": "В наличии",
            "instock_schema": "InStock"
        }
    ]
    saver.OUTPUT_DIR = tmp_path
    html_file = saver.render_html(data, filename="test.html")

    # Проверяем, что файл создан
    file_path = tmp_path / "test.html"
    assert file_path.exists()

    # Проверяем содержимое HTML
    content = file_path.read_text(encoding="utf-8")
    assert "<table" in content
    assert "Люстра" in content
    assert "12990" in content
    assert "https://example.com/lamp" in content
